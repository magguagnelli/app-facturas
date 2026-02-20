# core/excel_export.py
from __future__ import annotations

from datetime import datetime, date
from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _safe_str(v):
    # Evita fórmulas en Excel (CSV/Excel injection): si empieza con =, +, -, @ => prefix '
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def export_xlsx(
    rows: Sequence[dict],
    *,
    sheet_name: str = "Reporte",
    title: str | None = None,
    columns: list[tuple[str, str]] | None = None,
) -> bytes:
    """
    rows: lista de dicts.
    columns: opcional -> lista de (key, header). Si no se manda, usa llaves del primer row.
    Regresa bytes XLSX listo para enviar como respuesta.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita a 31 chars

    r = 1
    if title:
        ws.cell(row=r, column=1, value=title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(1, (len(columns) if columns else 1)))
        r += 2

    if not rows:
        ws.cell(row=r, column=1, value="Sin datos")
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # Determina columnas
    if columns is None:
        keys = list(rows[0].keys())
        columns = [(k, k) for k in keys]

    # Header
    for c, (_, header) in enumerate(columns, start=1):
        ws.cell(row=r, column=c, value=header)
    r += 1

    # Data
    for row in rows:
        for c, (key, _) in enumerate(columns, start=1):
            v = row.get(key)
            # normaliza fechas
            if isinstance(v, (datetime, date)):
                ws.cell(row=r, column=c, value=v)
            else:
                ws.cell(row=r, column=c, value=_safe_str(v))
        r += 1

    # Auto width (simple)
    for c in range(1, len(columns) + 1):
        col_letter = get_column_letter(c)
        max_len = 10
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            s = str(cell.value)
            max_len = max(max_len, len(s[:200]))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
