# services/facturas_listado_service.py
"""
Servicio para listado de facturas con paginación, filtros y exportación
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional
from datetime import datetime, date
from collections import namedtuple

import pandas as pd
import tempfile
import os

from core.db import get_conn


def _to_dict(row, cols):
    """Convierte una fila de cursor a diccionario"""
    if isinstance(row, dict):
        return row
    return dict(zip(cols, row))


def list_facturas_paginado(
    page: int = 1,
    per_page: int = 50,
    proveedor: Optional[str] = None,
    uuid: Optional[str] = None,
    area: Optional[int] = None,
    estatus_os: Optional[int] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Lista facturas con paginación y filtros múltiples.
   
    Returns:
        Dict con: items, total, page, per_page, total_pages
    """
    offset = (page - 1) * per_page
    params = []
   
    # Query base con TODOS los campos necesarios según factura_detalle.xlsx
    sql = """
            SELECT  
        --Campos de recepción
        row_number() OVER () AS "NO",
        u.nombre AS "RESPONSABLE DE CAPTURA A BASE",
        c.fecha_recepcion as "FECHA DE RECEPCION",
        a.nombre_area as "UNIDAD EJECUTORA DEL GASTO",
        
        --datos pago
        os.folio_oficio as "OFICIO",
        pr.rfc AS "RFC",
        pr.razon_social as "PROVEEDOR",
        os.cuenta_bancaria as "CUENTA BANCARIA",
        ct.num_contrato AS "CONTRATO",
        os.orden_suministro AS "ORDEN DE SUMINISTRO",
        os.folio_interno as "FOLIO INTERNO",
        c.uuid as "FOLIO FISCAL",
        os.validacion as "VALIDACION",
        os.mes_servicio as "MES DE SERVICIO",
        ct.ejercicio as "EJERCICIO FISCAL",
        os.monto_siniva as "MONTO SIN IVA",
        os.iva as "IVA",
        os.monto_c_iva AS "MONTO CON IVA",
        os.isr AS "ISR",
        os._5millar as "5 AL MILLAR",
        os.ieps AS "IEPS",
        os.re_imp_nomina AS "RETENCION IMPUESTO SOBRE LA NOMINA",
        os.riva as "RIVA",
        os.risr as "RISR",
        os.descuento AS "DESCUENTO",
        os.otras_contribuciones AS "OTRAS CONTRIBUCIONES",
        os.retenciones AS "RETENCION",
        os.penalizacion AS "PENALIZACION",
        os.deductiva AS "DEDUCTIVA",
        os.importe_pago AS "IMPORTE A PAGAR",
        os.importe_p_compromiso AS "IMPORTE PARA COMPROMISO",
        os.no_compromiso as "NO COMPROMISO",
        
        --ESTATUS
        eo.estatus_general "ESTATUS GENERAL",
        eo.estatus_reporte AS "ESTATUS REPORTE",    
        
        --DATOS PRESUPUESTALES
        p.capitulo AS "CAPITULO",
        p.partida_especifica AS "PARTIDA PRESUPUESTAL",
        p.pp AS "PROGRAMA PRESUPUESTAL",
        e.nombre as "ENTIDAD",
        e.id as "EF #",       
        
        --DATOS DE FISCALIZACION
        os.fecha_fiscalizacion as "FECHA DE FISCALIZACION",
        os.fiscalizador as "FISCALIZADOR",
        os.fecha_carga_sicop as "FECHA DE CARGA EN SICOP",
        os.responsable_carga_sicop as "RESPONSABLE DE CARGA SICOP",
        os.clc AS "CLC 2024",
        os.numero_solicitud_pago AS "NUMERO DE SOLICITUD DE PAGO 2024",
        os.clc25 AS "CLC 2025",
        os.numero_solicitud_pago25 AS "NUMERO DE SOLICITUD DE PAGO 2025",
        os.clc26 AS "CLC 2026",
        os.numero_solicitud_pago26 AS "NUMERO DE SOLICITUD DE PAGO 2026",
        os.clc27 AS "CLC 2027",
        os.numero_solicitud_pago27 AS "NUMERO DE SOLICITUD DE PAGO 2027",
        os.estatus_siaff as "ESTATUS SIAFF",
        os.fecha_pago AS "FECHA DE PAGO",
        
        --devolucion
        os.oficio_dev AS "OFICIO DEV",
        os.fecha_dev AS "FECHA DEV",
        os.motivo_dev AS "MOTIVO DEV",    
        
        --final
        os.observaciones as observaciones_os,
        os.responsable_fis as "RESPONSABLE DOC FIS",
        
        --nuevos
        os.fecha_pr AS "FECHA DE PAGO REFERENCIADO",
        os.inmueble AS "INMUEBLE",
        os.periodo AS "PERIODO",
        os.recargos AS "RECARGARGOS EN PAGO DE SERVICIOS",
        os.observacion_pr AS "OBSERVACION",
        os.corte_presupuesto AS "CORTE PRESUPUESTO",
        os.fecha_turno AS "FECHA DE TURNO",
        
        --campos sin uso definido
        ct.rfc_pp,
        ct.f_inicio as contrato_f_inicio,
        ct.f_fin as contrato_f_fin,
        ct.mes as contrato_mes,
        ct.monto_total as contrato_monto_total,
        ct.monto_maximo as contrato_monto_maximo,
        ct.monto_ejercido as contrato_monto_ejercido,
        ct.saldo_disponible as contrato_saldo_disponible,
        ct.estatus as contrato_estatus, 
        a.desc_area,
        os.fecha_orden,
        c.rfc_emisor,
        pr.tipo_persona as proveedor_tipo,
        eo.estado_resumen,
        p.des_cap,
        p.concepto,
        p.des_concepto,
        p.uso_partida,
        p.des_uso_partida,
        p.des_pe,
        p.tipo_gasto,
        p.austeridad,    
        p.des_pp,
        p.monto_total as partida_monto_total,
        p.observaciones as partida_observaciones,
        c.fecha_emision,
        c.monto_total,
        c.estatus as cfdi_estatus,
        c.onservaciones as observaciones_cfdi,
        os.fecha_factura,
        os.estatus as estatus_os,
        os.solicitud,
        --ids
        c.id as cfdi_id,
        a.id as area_id,
        os.id as os_id,
        ct.id as contrato_id,
        pr.id as proveedor_id,
        eo.id as estado_orden_id,
        p.id as partida_id
    FROM cat_facturas.cfdi c
    INNER JOIN cat_facturas.orden_suministro os ON os.id = c.orden_suministro
    LEFT JOIN cat_facturas.proveedor pr ON pr.id = os.proveedor
    LEFT JOIN cat_facturas.partida p ON p.id = os.partida
    LEFT JOIN cat_facturas.contrato ct ON ct.id = p.contrato
    LEFT JOIN cat_facturas.area a ON a.id = ct.area
    LEFT JOIN cat_facturas.entidad e ON e.id = p.entidad
    LEFT JOIN cat_facturas.estado_orden eo ON eo.id = os.estatus
    left join cat_facturas.usuario u on c.resp_captura= u.correo
    """
   
    # Aplicar filtros
    if proveedor:
        sql += " AND (pr.rfc ILIKE %s OR pr.razon_social ILIKE %s)"
        like_prov = f"%{proveedor}%"
        params.extend([like_prov, like_prov])
   
    if uuid:
        sql += " AND c.uuid ILIKE %s"
        params.append(f"%{uuid}%")
   
    if area:
        sql += " AND a.id = %s"
        params.append(area)
   
    if estatus_os:
        sql += " AND os.estatus = %s"
        params.append(estatus_os)
   
    if fecha_inicio:
        sql += " AND c.fecha_recepcion >= %s"
        params.append(fecha_inicio)
   
    if fecha_fin:
        sql += " AND c.fecha_recepcion <= %s"
        params.append(fecha_fin)
   
    # Contar total de registros
    count_sql = f"SELECT COUNT(*) FROM ({sql}) AS subq"
   
    with get_conn() as conn:
        with conn.cursor() as cur:
            # Total count
            cur.execute(count_sql, params)
            #print(cur.fetchall())
            total = cur.fetchone()["count"]
            #print(total["count"])
           
            # Datos paginados
            sql += " ORDER BY c.fecha_recepcion DESC, c.id DESC"
            sql += f" LIMIT %s OFFSET %s"
            params.extend([per_page, offset])
           
            cur.execute(sql, params)
            cols = [d[0] for d in cur.description]
            #print("COLUMNAS DEL QUERY:")
            #print(cols)
            items = [_to_dict(row, cols) for row in cur.fetchall()]
   
    total_pages = (total + per_page - 1) // per_page
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
    }


def exportar_facturas_excel(
    proveedor: Optional[str] = None,
    uuid: Optional[str] = None,
    area: Optional[int] = None,
    estatus_os: Optional[int] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> str:
    """
    Exporta todas las facturas (con filtros opcionales) a Excel.
    Retorna la ruta del archivo temporal.
    """
    # Obtener TODAS las facturas sin paginación
    result = list_facturas_paginado(
        page=1,
        per_page=999999,
        proveedor=proveedor,
        uuid=uuid,
        area=area,
        estatus_os=estatus_os,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )
    
    items = result["items"]
    
    if not items:
        return None
    
    # Convertir a DataFrame
    df = pd.DataFrame(items)
    
    # Columnas ordenadas
    columnas_ordenadas = [
    "NO", "RESPONSABLE DE CAPTURA A BASE", "FECHA DE RECEPCION", "UNIDAD EJECUTORA DEL GASTO",
    "OFICIO", "RFC", "PROVEEDOR", "CUENTA BANCARIA", "CONTRATO",
    "ORDEN DE SUMINISTRO", "FOLIO INTERNO", "FOLIO FISCAL", "VALIDACION", "MES DE SERVICIO",
    "EJERCICIO FISCAL", "MONTO SIN IVA", "IVA", "MONTO CON IVA", "ISR",
    "5 AL MILLAR", "IEPS", "RETENCION IMPUESTO SOBRE LA NOMINA", "RIVA", "RISR",
    "DESCUENTO", "OTRAS CONTRIBUCIONES", "RETENCIONES", "PENALIZACION", "DEDUCTIVA",
    "IMPORTE A PAGAR", "IMPORTE PARA COMPROMISO", "NO COMPROMISO","ESTATUS GENERAL", "ESTATUS REPORTE",
    "CAPITULO", "PARTIDA PRESUPUESTAL", "PROGRAMA PRESUPUESTAL","ENTIDAD", "EF #",
    "FECHA DE FISCALIZACION", "FISCALIZADOR", "FECHA DE CARGA EN SICOP", "RESPONSABLE DE CARGA SICOP",
    "CLC 2024", "NUMERO DE SOLICITUD DE PAGO 2024", 
    "CLC 2025", "NUMERO DE SOLICITUD DE PAGO 2025",
    "CLC 2026", "NUMERO DE SOLICITUD DE PAGO 2026",
    "CLC 2027","NUMERO DE SOLICITUD DE PAGO 2027",
    "ESTATUS SIAFF", "FECHA DE PAGO", 
    "OFICIO DEV", "FECHA DEV", "MOTIVO DEV",
    "FECHA DE PAGO REFERENCIADO", "INMUEBLE", "PERIODO", "RECARGARGOS EN PAGO DE SERVICIOS", "OBSERVACIONES_PR",
    "CORTE PRESUPUESTO", "RESPONSABLE DOC FIS", "FECHA DE TURNO"
]

    
    # Filtrar solo columnas que existen
    columnas_disponibles = [col for col in columnas_ordenadas if col in df.columns]
    df = df[columnas_disponibles]
    
    # Crear archivo temporal
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    filepath = temp_file.name
    temp_file.close()
    
    # Escribir Excel con ajuste automático de columnas
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Facturas', index=False)
        
        # Ajustar ancho de columnas
        worksheet = writer.sheets['Facturas']
        
        from openpyxl.utils import get_column_letter
        
        for idx, col in enumerate(df.columns, 1):
            try:
                # Calcular longitud máxima del contenido
                col_series = df[col].astype(str)
                max_content_length = col_series.str.len().max()
                
                # Comparar con longitud del header
                max_length = max(
                    max_content_length if pd.notna(max_content_length) else 0,
                    len(str(col))
                )
                
                # Ajustar ancho (mínimo 10, máximo 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                
                # Aplicar ancho
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = adjusted_width
                
            except Exception as e:
                # Si falla, usar ancho por defecto de 15
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = 15
    
    return filepath


def get_filtros_opciones() -> Dict[str, List[Dict]]:
    """
    Retorna las opciones disponibles para los filtros.
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            # Áreas
            cur.execute("""
                SELECT id, nombre_area 
                FROM cat_facturas.area 
                WHERE estatus = 'ACTIVO'
                ORDER BY nombre_area
            """)
            areas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]
            
            # Estados de orden
            cur.execute("""
                SELECT id, estatus_general, estatus_reporte
                FROM cat_facturas.estado_orden
                ORDER BY id
            """)
            estados_orden = [
                {
                    "id": r[0],
                    "estatus_general": r[1],
                    "estatus_reporte": r[2]
                }
                for r in cur.fetchall()
            ]
    
    return {
        "areas": areas,
        "estados_orden": estados_orden,
    }
